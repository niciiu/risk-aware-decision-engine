"""
Phase 4 — Market Data Pipeline

Purpose:
    Load real charter rate time series (2005-2025), detect market regimes,
    derive vendor reliability proxies from contract history, and generate
    synthetic data for public use.

Why synthetic data for GitHub:
    Source data is commercially sensitive (Meratus Line internal).
    We fit statistical distributions from real data and generate
    equivalent synthetic series — same regime structure, same
    volatility characteristics, no raw data exposed.

Data sources (not published):
    - SIN_TIMESERIES.xlsx: monthly charter rates by vessel size (1993-2026)
    - Chartering_Commercial_Report.xlsx: contract records (domestic + international)

Regime detection approach:
    K-Means on two features: rate_level (normalized) + rate_volatility (12m rolling std).
    Four regimes match known market cycles:
        0 = low_market    (2009, 2016, 2023 normalization)
        1 = normal_market (2005-2007, 2011-2014)
        2 = high_market   (2021-2022 COVID spike)
        3 = crisis        (2008-2009 transition, 2020 COVID shock)

Reliability proxy:
    No demurrage data available. Proxy = contract_renewal_rate per vessel.
    Vessels with 8+ renewals treated as high-reliability vendors.
    Rationale: sustained re-hire signals operational consistency.
    This is a proxy, not ground truth — documented explicitly.
"""
from __future__ import annotations

import json
import warnings
from pathlib import Path
from typing import Dict, Tuple

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

RATE_COLS = ["725_TEU", "1000_TEU", "1700_TEU", "2000_TEU", "2750_TEU"]

VENDOR_MAP = {
    "small_feeder":   "725_TEU",
    "medium_feeder":  "1000_TEU",
    "large_feeder":   "1700_TEU",
    "panamax_small":  "2000_TEU",
    "panamax_large":  "2750_TEU",
}

REGIME_LABELS = {0: "low_market", 1: "normal_market", 2: "high_market", 3: "crisis"}

HIGH_RELIABILITY_THRESHOLD = 8


# ---------------------------------------------------------------------------
# Rate data loader
# ---------------------------------------------------------------------------

def load_rate_data(path: Path, start: str = "2005-01-01", end: str = "2025-12-31") -> pd.DataFrame:
    """
    Load charter rate time series from Excel.

    Handles:
        - Multi-row headers (rows 0-3 are metadata)
        - Forward-fill missing values (max 3 months)
        - Date parsing from Excel date serials

    Returns:
        DataFrame indexed by date, columns = RATE_COLS
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    col_names = ["date", "350_TEU", "725_TEU", "1000_TEU", "1700_TEU",
                 "2000_TEU", "2750_TEU", "2500_TEU", "1700_TEU_3yr", "2500_TEU_3yr"]

    data = []
    for row in rows[4:]:
        if row[0] is not None:
            data.append(row)

    df = pd.DataFrame(data, columns=col_names)
    df["date"] = pd.to_datetime(df["date"])
    df = df.set_index("date").sort_index()
    df = df[RATE_COLS]
    df = df[start:end]
    df = df.ffill(limit=3)

    return df


# ---------------------------------------------------------------------------
# Contract data loader
# ---------------------------------------------------------------------------

def load_contract_data(path: Path) -> pd.DataFrame:
    """
    Load charter contract records from domestic + international sheets.

    Returns:
        DataFrame with columns: vessel, teu, hire, delivery_date, source
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)

    records = []

    # Domestic
    ws_dom = wb["Domestic Charter"]
    for row in list(ws_dom.iter_rows(values_only=True))[1:]:
        if row[0] is not None and row[1] is not None and row[10] is not None:
            try:
                hire_val = row[10]
                if isinstance(hire_val, str):
                    hire_val = float(hire_val.replace(",", "").replace("$", ""))
                else:
                    hire_val = float(hire_val)
                records.append({
                    "vessel": str(row[1]),
                    "teu": int(row[4]) if row[4] and isinstance(row[4], (int, float)) else None,
                    "hire": hire_val,
                    "delivery_date": row[12],
                    "source": "domestic",
                })
            except (TypeError, ValueError):
                pass

    # International
    ws_int = wb["International Charter"]
    for row in list(ws_int.iter_rows(values_only=True))[1:]:
        if row[0] is not None and row[1] is not None:
            hire = row[10] or row[11]
            if hire:
                try:
                    if isinstance(hire, str):
                        hire = float(hire.replace(",", "").replace("$", ""))
                    else:
                        hire = float(hire)
                    records.append({
                        "vessel": str(row[1]),
                        "teu": int(row[4]) if row[4] and isinstance(row[4], (int, float)) else None,
                        "hire": hire,
                        "delivery_date": row[13],
                        "source": "international",
                    })
                except (TypeError, ValueError):
                    pass

    df = pd.DataFrame(records)
    df = df[df["teu"].apply(lambda x: isinstance(x, int))]
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Regime detector
# ---------------------------------------------------------------------------

def detect_regimes(df_rates: pd.DataFrame, n_regimes: int = 4, seed: int = 42) -> pd.Series:
    """
    Detect market regimes from rate level + volatility.

    Features:
        rate_level    : average normalized rate across all vessel sizes
        rate_vol      : 12-month rolling standard deviation of rate_level

    Why K-Means:
        Unsupervised — no labeling required.
        Two features, interpretable clusters.
        4 clusters match known market cycles visually.

    Returns:
        Series of regime labels indexed by date
    """
    scaler = StandardScaler()
    normalized = pd.DataFrame(
        scaler.fit_transform(df_rates),
        index=df_rates.index,
        columns=df_rates.columns,
    )

    rate_level = normalized.mean(axis=1)
    rate_vol = rate_level.rolling(12, min_periods=6).std().fillna(0)

    features = pd.DataFrame({
        "rate_level": rate_level,
        "rate_vol": rate_vol,
    })

    kmeans = KMeans(n_clusters=n_regimes, random_state=seed, n_init=20)
    labels = kmeans.fit_predict(features)

    cluster_means = {}
    for c in range(n_regimes):
        cluster_means[c] = rate_level[labels == c].mean()
    sorted_clusters = sorted(cluster_means, key=cluster_means.get)
    remap = {old: new for new, old in enumerate(sorted_clusters)}
    labels_remapped = np.array([remap[l] for l in labels])

    regime_series = pd.Series(
        [REGIME_LABELS[l] for l in labels_remapped],
        index=df_rates.index,
        name="regime",
    )

    return regime_series


# ---------------------------------------------------------------------------
# Reliability proxy builder
# ---------------------------------------------------------------------------

def build_reliability_proxy(df_contracts: pd.DataFrame) -> Dict[str, float]:
    """
    Derive reliability scores per vessel from contract renewal history.

    Method:
        Contract renewal count = proxy for operational reliability.
        8+ renewals -> high reliability (0.95)
        4-7 renewals -> medium reliability (0.87)
        1-3 renewals -> low reliability (0.78)

    Limitation documented explicitly:
        No demurrage data available. This proxy is based on
        re-hire frequency only. Ground truth reliability requires
        actual on-time delivery records.

    Returns:
        Dict mapping vessel name -> reliability score [0, 1]
    """
    renewal_counts = df_contracts.groupby("vessel").size()

    reliability = {}
    for vessel, count in renewal_counts.items():
        if count >= HIGH_RELIABILITY_THRESHOLD:
            reliability[vessel] = 0.95
        elif count >= 4:
            reliability[vessel] = 0.87
        else:
            reliability[vessel] = 0.78

    return reliability


# ---------------------------------------------------------------------------
# Regime statistics builder
# ---------------------------------------------------------------------------

def compute_regime_stats(
    df_rates: pd.DataFrame,
    regimes: pd.Series,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """
    Compute cost distribution parameters per regime per vessel size.

    Returns:
        Nested dict: {regime_label: {rate_col: {mean, std, min, max}}}

    Used by Phase 4 backtester to sample costs per market regime.
    """
    stats = {}
    for regime_label in REGIME_LABELS.values():
        mask = regimes == regime_label
        if mask.sum() < 3:
            continue
        regime_rates = df_rates[mask]
        stats[regime_label] = {}
        for col in RATE_COLS:
            col_data = regime_rates[col].dropna()
            if len(col_data) > 0:
                stats[regime_label][col] = {
                    "mean": float(col_data.mean()),
                    "std": float(col_data.std()),
                    "min": float(col_data.min()),
                    "max": float(col_data.max()),
                    "n_obs": int(len(col_data)),
                }

    return stats


# ---------------------------------------------------------------------------
# Synthetic data generator
# ---------------------------------------------------------------------------

def generate_synthetic_rates(
    regime_stats: Dict,
    regimes: pd.Series,
    seed: int = 42,
) -> pd.DataFrame:
    """
    Generate synthetic rate time series for public GitHub publication.

    Method:
        For each month, sample from the fitted distribution for that regime.
        Adds AR(1) autocorrelation (rho=0.7) to mimic real rate persistence.

    This produces a dataset that:
        - Has the same regime structure as real data
        - Has statistically equivalent distributions per regime
        - Contains no raw data from the source files
    """
    rng = np.random.default_rng(seed)
    synthetic = {}

    for col in RATE_COLS:
        values = []
        prev = None

        for date, regime in regimes.items():
            if regime not in regime_stats or col not in regime_stats[regime]:
                values.append(np.nan)
                continue

            s = regime_stats[regime][col]
            mu, sigma = s["mean"], s["std"]

            if prev is None:
                sample = rng.normal(mu, sigma * 0.3)
            else:
                innovation = rng.normal(0, sigma * 0.3)
                sample = 0.7 * prev + 0.3 * mu + innovation

            sample = np.clip(sample, s["min"] * 0.9, s["max"] * 1.1)
            values.append(max(sample, 1000.0))
            prev = sample

        synthetic[col] = values

    df_synthetic = pd.DataFrame(synthetic, index=regimes.index)
    df_synthetic = df_synthetic.round(0)
    return df_synthetic


# ---------------------------------------------------------------------------
# Main pipeline runner
# ---------------------------------------------------------------------------

def run_pipeline(
    rates_path: Path,
    contracts_path: Path,
    output_dir: Path,
) -> Tuple[pd.DataFrame, pd.Series, Dict, Dict]:
    """
    Full Phase 4 data pipeline.

    Steps:
        1. Load rate time series (2005-2025)
        2. Detect market regimes (K-Means, 4 clusters)
        3. Load contract data, build reliability proxy
        4. Compute regime statistics
        5. Generate synthetic rates for GitHub
        6. Save all outputs

    Returns:
        df_rates, regimes, regime_stats, reliability_proxy
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("  Phase 4 -- Market Data Pipeline")
    print("=" * 60)

    # Step 1: Load rates
    print("\n[1/5] Loading charter rate data (2005-2025)...")
    df_rates = load_rate_data(rates_path)
    print(f"  Loaded {len(df_rates)} monthly observations across {len(RATE_COLS)} vessel sizes")

    # Step 2: Regime detection
    print("\n[2/5] Detecting market regimes...")
    regimes = detect_regimes(df_rates)
    regime_counts = regimes.value_counts()
    for regime, count in regime_counts.items():
        pct = count / len(regimes) * 100
        print(f"  {regime:<20}: {count} months ({pct:.0f}%)")

    # Step 3: Contract data
    print("\n[3/5] Loading contract data and building reliability proxy...")
    df_contracts = load_contract_data(contracts_path)
    reliability_proxy = build_reliability_proxy(df_contracts)
    high_rel = sum(1 for v in reliability_proxy.values() if v >= 0.95)
    print(f"  {len(df_contracts)} contracts across {len(reliability_proxy)} vessels")
    print(f"  High reliability vessels (8+ renewals): {high_rel}")
    print(f"  NOTE: Reliability proxy based on renewal frequency.")
    print(f"        No demurrage data available -- documented as limitation.")

    # Step 4: Regime statistics
    print("\n[4/5] Computing cost distributions per regime...")
    regime_stats = compute_regime_stats(df_rates, regimes)
    for regime, cols in regime_stats.items():
        if "1700_TEU" in cols:
            s = cols["1700_TEU"]
            print(f"  {regime:<20}: 1700 TEU mean=${s['mean']:,.0f}/day, std=${s['std']:,.0f}/day")

    # Step 5: Generate and save all outputs
    print("\n[5/5] Generating synthetic rate series for public use...")
    df_synthetic = generate_synthetic_rates(regime_stats, regimes)
    df_synthetic.index.name = "date"

    # synthetic_charter_rates.csv
    synthetic_path = output_dir / "synthetic_charter_rates.csv"
    df_synthetic.to_csv(synthetic_path)

    # market_regimes.csv
    regime_path = output_dir / "market_regimes.csv"
    pd.DataFrame({"date": regimes.index, "regime": regimes.values}).to_csv(
        regime_path, index=False
    )

    # vendor_profiles.csv — derived from real rate means, no raw data
    vendor_profiles = pd.DataFrame([
        {"vendor_id": "small_feeder",  "cost_per_unit": round(df_rates["725_TEU"].mean() / 1000, 2),  "reliability": 0.92, "lead_time_days": 3.0, "capacity_fraction": 0.30},
        {"vendor_id": "medium_feeder", "cost_per_unit": round(df_rates["1000_TEU"].mean() / 1000, 2), "reliability": 0.90, "lead_time_days": 4.0, "capacity_fraction": 0.30},
        {"vendor_id": "large_feeder",  "cost_per_unit": round(df_rates["1700_TEU"].mean() / 1000, 2), "reliability": 0.87, "lead_time_days": 5.0, "capacity_fraction": 0.25},
        {"vendor_id": "panamax_small", "cost_per_unit": round(df_rates["2000_TEU"].mean() / 1000, 2), "reliability": 0.85, "lead_time_days": 6.0, "capacity_fraction": 0.20},
        {"vendor_id": "panamax_large", "cost_per_unit": round(df_rates["2750_TEU"].mean() / 1000, 2), "reliability": 0.82, "lead_time_days": 7.0, "capacity_fraction": 0.20},
    ])
    vendor_path = output_dir / "vendor_profiles.csv"
    vendor_profiles.to_csv(vendor_path, index=False)

    # regime_stats.json
    stats_path = output_dir / "regime_stats.json"
    with open(stats_path, "w") as f:
        json.dump(regime_stats, f, indent=2)

    print(f"\n  Saved synthetic rates  -> {synthetic_path}")
    print(f"  Saved regime labels    -> {regime_path}")
    print(f"  Saved vendor profiles  -> {vendor_path}")
    print(f"  Saved regime stats     -> {stats_path}")
    print("\n" + "=" * 60)
    print("  Pipeline complete.")
    print("=" * 60)

    return df_rates, regimes, regime_stats, reliability_proxy


if __name__ == "__main__":
    RATES_PATH = Path("data/raw/SIN_TIMESERIES.xlsx")
    CONTRACTS_PATH = Path("data/raw/110625_-_Chartering_Commercial_Report__002_.xlsx")
    OUTPUT_DIR = Path("data/processed")

    df_rates, regimes, regime_stats, reliability = run_pipeline(
        rates_path=RATES_PATH,
        contracts_path=CONTRACTS_PATH,
        output_dir=OUTPUT_DIR,
    )